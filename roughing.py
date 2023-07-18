# # import io
# # import pdfminer.high_level
# # from pdfminer.high_level import extract_text
# # from pdfminer.layout import LAParams
# # from openpyxl import Workbook
# #
# # # Define the dictionary
# # my_dict = {
# #     'EMPLOYMENT AUTHORIZATIONS': [7],
# #     'TYPE': [11, 38, 44],
# #     'FULL/PART-TIME': [16],
# #     'STATUS': [19],
# #     'START DATE': [26, 53, 64],
# #     'END DATE': [31, 56, 67],
# #     'EMPLOYER INFORMATION': [36],
# #     'EMPLOYER NAME': [41, 47],
# #     'AUTHORIZATION DATES': [50, 61],
# #     'CITY & STATE': [58, 70]
# # }
# #
# # # Define the order in which the lines should be printed
# # order = [
# #     ('TYPE', 'FULL/PART-TIME'),
# #     ('FULL/PART-TIME', 'STATUS'),
# #     ('STATUS', 'START DATE'),
# #     ('START DATE', 'END DATE'),
# #     ('EMPLOYER INFORMATION', 'AUTHORIZATION DATES'),
# #     ('AUTHORIZATION DATES', 'EMPLOYER NAME'),
# #     ('EMPLOYER NAME', 'START DATE'),
# #     ('START DATE', 'END DATE'),
# #     ('END DATE', 'CITY & STATE'),
# #     ('CITY & STATE', 'TYPE'),
# #     ('TYPE', 'AUTHORIZATION DATES'),
# #     ('AUTHORIZATION DATES', 'EMPLOYER NAME'),
# #     ('EMPLOYER NAME', 'START DATE'),
# #     ('START DATE', 'END DATE'),
# #     ('END DATE', 'CITY & STATE'),
# #     ('CITY & STATE', '')  # Include the last line of the file
# # ]
# #
# # # Extract text from the PDF file
# # with open('requested.pdf', 'rb') as file:
# #     buffer = io.BytesIO(file.read())
# # text = extract_text(buffer, laparams=LAParams(), page_numbers=[1])
# #
# # # Split the text into lines
# # lines = text.split('\n')
# #
# # # Create an Excel workbook and sheet
# # workbook = Workbook()
# # sheet = workbook.active
# #
# # # Write the headers to the sheet
# # headers = [' '.join([start_key, 'to', end_key]) for start_key, end_key in order]
# # sheet.append(headers)
# #
# # # Write the data to the sheet
# # for start_key, end_key in order:
# #     start = my_dict[start_key][0] + 1
# #     if end_key:
# #         end = my_dict[end_key][0]
# #     else:
# #         end = len(lines)
# #     data = [line.strip() for line in lines[start:end]]
# #     sheet.append(data)
# #
# # # Save the Excel file
# # workbook.save('output.xlsx')
# #
# # import io
# # import pdfminer.high_level
# # from pdfminer.high_level import extract_text
# # from pdfminer.layout import LAParams
# #
# # # Define the dictionary
# # my_dict = {
# #     'EMPLOYMENT AUTHORIZATIONS': [7],
# #     'TYPE': [11, 38, 44],
# #     'FULL/PART-TIME': [16],
# #     'STATUS': [19],
# #     'START DATE': [26, 53, 64],
# #     'END DATE': [31, 56, 67],
# #     'EMPLOYER INFORMATION': [36],
# #     'EMPLOYER NAME': [41, 47],
# #     'AUTHORIZATION DATES': [50, 61],
# #     'CITY & STATE': [58, 70]
# # }
# #
# # # Define the order in which the lines should be printed
# # order = [
# #     ('TYPE', 'FULL/PART-TIME'),
# #     ('FULL/PART-TIME', 'STATUS'),
# #     ('STATUS', 'START DATE'),
# #     ('START DATE', 'END DATE'),
# #     ('EMPLOYER INFORMATION', 'TYPE'),
# #     ('TYPE', 'EMPLOYER NAME'),
# #     ('EMPLOYER NAME', 'AUTHORIZATION DATES'),
# #     ('START DATE', 'END DATE'),
# #     ('END DATE', 'CITY & STATE'),
# #     ('CITY & STATE', 'TYPE'),
# #     ('TYPE', 'AUTHORIZATION DATES'),
# #     ('AUTHORIZATION DATES', 'EMPLOYER NAME'),
# #     ('EMPLOYER NAME', 'START DATE'),
# #     ('START DATE', 'END DATE'),
# #     ('END DATE', 'CITY & STATE'),
# #     ('CITY & STATE', '')  # Include the last line of the file
# # ]
# #
# # # Extract text from the PDF file
# # with open('requested.pdf', 'rb') as file:
# #     buffer = io.BytesIO(file.read())
# # text = extract_text(buffer, laparams=LAParams(), page_numbers=[1])
# #
# # # Split the text into lines
# # lines = text.split('\n')
# #
# # # Create an empty dictionary
# # data_dict = {}
# #
# # # Extract the data and store it in the dictionary
# # for start_key, end_key in order:
# #     start = my_dict[start_key][0] + 1
# #     if end_key:
# #         end = my_dict[end_key][0]
# #     else:
# #         end = len(lines)
# #     data = [line.strip() for line in lines[start:end]]
# #     key = f"{start_key} to {end_key}" if end_key else start_key
# #     data_dict[key] = data
# #
# # # Print the data dictionary
# # print(data_dict)
#
#
# # import io
# # import pdfminer.high_level
# # from pdfminer.high_level import extract_text
# # from pdfminer.layout import LAParams
# #
# # # Define the dictionary
# # my_dict = {
# #     'EMPLOYMENT AUTHORIZATIONS': [7],
# #     'TYPE': [11, 38, 44],
# #     'FULL/PART-TIME': [16],
# #     'STATUS': [19],
# #     'START DATE': [26, 53, 64],
# #     'END DATE': [31, 56, 67],
# #     'EMPLOYER INFORMATION': [36],
# #     'EMPLOYER NAME': [41, 47],
# #     'AUTHORIZATION DATES': [50, 61],
# #     'CITY & STATE': [58, 70]
# # }
# #
# #
# # # Get a list of keys in the order they appear in the my_dict dictionary
# # keys = []
# # for key in my_dict:
# #     index = my_dict[key][0]
# #     keys.insert(index, key)
# #
# # # Use the keys to update the order list
# # order = [(keys[i], keys[i+1]) for i in range(len(keys)-1)]
# # order.append((keys[-1], ''))  # Include the last line of the file
# #
# # # Extract text from the PDF file
# # with open('requested.pdf', 'rb') as file:
# #     buffer = io.BytesIO(file.read())
# # text = extract_text(buffer, laparams=LAParams(), page_numbers=[1])
# #
# # # Split the text into lines
# # lines = text.split('\n')
# #
# # # Create an empty dictionary
# # data_dict = {}
# #
# # # Extract the data and store it in the dictionary
# # for start_key, end_key in order:
# #     start = my_dict[start_key][0] + 1
# #     if end_key:
# #         end = my_dict[end_key][0]
# #     else:
# #         end = len(lines)
# #     data = [line.strip() for line in lines[start:end]]
# #     key = f"{start_key} to {end_key}" if end_key else start_key
# #     data_dict[key] = data
# #
# # # Print the data dictionary
# # print(data_dict)
#
#
import io
from pdfminer.high_level import extract_text
from pdfminer.layout import LAParams
import openpyxl

# Define the dictionary
my_dict = {
    'EMPLOYMENT AUTHORIZATIONS': [7],
    'TYPE': [11, 38, 44],
    'FULL/PART-TIME': [16],
    'STATUS': [19],
    'START DATE': [26, 53, 64],
    'END DATE': [31, 56, 67],
    'EMPLOYER INFORMATION': [36],
    'EMPLOYER NAME': [41, 47],
    'AUTHORIZATION DATES': [50, 61],
    'CITY & STATE': [58, 70],
    'CHANGE OF STATUS/CAP-GAP EXTENSION':[73]
}

# Define the order in which the lines should be printed
order = [
    ('EMPLOYMENT AUTHORIZATIONS', 'TYPE'),
    ('TYPE', 'FULL/PART-TIME[0]'),
    ('FULL/PART-TIME[0]','STATUS[0]'),
    ('STATUS[0]','START DATE[0]'),
    ('START DATE[0]','END DATE[0]'),
    ('END DATE[0]','EMPLOYER INFORMATION[0]'),
    ('TYPE[1]', 'EMPLOYER NAME[0]'),
    ('EMPLOYER NAME[0]', 'TYPE[2]'),
    ('TYPE[2]', 'EMPLOYER NAME[1]'),
    ('EMPLOYER NAME[1]', 'AUTHORIZATION DATES[0]'),
    ('AUTHORIZATION DATES[0]', 'START DATE[1]'),
    ('START DATE[1]', 'END DATE[1]'),
    ('END DATE[1]', 'CITY & STATE[0]'),
    ('CITY & STATE[0]', 'AUTHORIZATION DATES[1]'),
    ('AUTHORIZATION DATES[1]', 'START DATE[2]'),
    ('START DATE[2]', 'END DATE[2]'),
    ('END DATE[2]', 'CITY & STATE[1]'),
    ('CITY & STATE[1]', 'CHANGE OF STATUS/CAP-GAP EXTENSION[0]')  # Include the last line of the file
]

pdf_path = 'requested.pdf'

# Extract the text from the PDF file and split it into lines
with open(pdf_path, 'rb') as pdf_file:
    text = extract_text(pdf_file, laparams=LAParams(),page_numbers=[1])
    lines = text.split('\n')

# Create a dictionary to store the index of each key
key_index = {}
for key in my_dict:
    key_index[key] = 0

# Extract the data and store it in the dictionary
data_dict = {}
workbook = openpyxl.Workbook()
sheet = workbook.active
for start_key, end_key in order:
    # Extract the start and end indices from the dictionary, handling keys with indices
    if '[' in start_key:
        start_key, start_index = start_key.split('[')
        start_index = int(start_index.strip(']'))   # Convert to 0-indexed
        #print(start_key,start_index)
    else:
        start_index = key_index[start_key]

    if end_key is None:
        end_index = len(lines)
    else:
        if '[' in end_key:
            end_key, end_index = end_key.split('[')
            end_index = int(end_index.strip(']'))   # Convert to 0-indexed
            #print(end_key,end_index)
        else:
            end_index = key_index.get(end_key, len(my_dict[end_key]))

    # Extract the data and store it in the data dictionary

    data = [lines[i].strip() for i in range(my_dict[start_key][start_index], my_dict[end_key][end_index])]
    key = f"{start_key} to {end_key}" if end_key else start_key
    print(data)
    # print(key)


    data_dict[key] = data
    print(data_dict)


    # Increment the start index in the key index dictionary
    key_index[start_key] = start_index + 1

    # # Create a new Excel workbook and select the active sheet

    #
    # Write the data dictionary to the sheet
    for column, (key, data) in enumerate(data_dict.items(), start=1):
        # Write the key as the header
        sheet.cell(row=1, column=column, value=key)
        # Write the data to the corresponding column
        for row, value in enumerate(data[1:], start=2):
            sheet.cell(row=row, column=column, value=value)

    # Save the workbook to a file
    workbook.save('data.xlsx')

# Print the data dictionary
#print(data_dict)
print(data_dict['TYPE to FULL/PART-TIME'][1])
print(data_dict['START DATE to END DATE'][1])


