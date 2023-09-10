import time
import re
import openpyxl
from preProcessor.AddIndividual import AddIndividual
import pandas as pd
import configparser
from selenium.webdriver.support.ui import *
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from preProcessor.issmfilelog import logger
from selenium.webdriver.support import expected_conditions as ec
from flask_socketio import SocketIO
from flask import Flask
from flask_cors import CORS
import secrets
import math

app = Flask(__name__, template_folder='../../', static_folder='../../static')
CORS(app)
app.secret_key = secrets.token_bytes(32)
socketio = SocketIO(app, cors_allowed_origins="*")

@socketio.on('connect')
def handle_connect():
    print('Client connected')
@socketio.on('disconnect')
def handle_disconnect():
    print('Client disconnected')
class ProgressBar:
    def __init__(self, max_count):
        self.processed_count = 0
        self.max_count = max_count
        self.success_count = 0
        self.failure_count = 0
        self.deferral_count = 0

    def getProcessCount(self):
        return self.processed_count

    def __str__(self):
        return f"ProgressBar(processed_count={self.processed_count}, success_count={self.success_count}, failure_count={self.failure_count}, deferral_count={self.deferral_count}, max_count={self.max_count})"

def adding_to_excel(student, driver, config, progress_bar):
    logger.info(f"inside adding_to_excel function for {student.CampusID}")
    try:
        error_element = driver.find_element(By.ID, config['ID_XPATH']['error_element'])
        error_text = error_element.text
        if 'Individual With Duplicate Campus ID Found - Please Correct' in error_text:
            logger.info(f'Error found : {error_text} for Last Name {student.Surname}')
        workbook = load_workbook("preProcessor/Duplicate.xlsx")
        sheet = workbook.active
        if "Birthdate" not in [cell.value for cell in sheet[1]]:
            # Append the header if it is not present
            sheet.append(["Campus ID", "Admissions ID", "Name", "Birthdate"])
        sheet.append([student.CampusID, student.AdmissionsID, student.GivenName, student.Birthdate])
        workbook.save("preProcessor/Duplicate.xlsx")
        logger.info(f"added student successfully adding_to_excel function for {student.CampusID}")
        progress_bar.deferral_count += 1
        return True
    except Exception as e:
        logger.error(f"adding student failed adding_to_excel function for {student.CampusID}")
        logger.error(e)
        return False
def duplicate_steps(student, driver, domain_url, config):
    logger.info(f"inside duplicate_steps function for {student.CampusID}")
    # time.sleep(1)
    try:
        logger.info(f"starting removal of student id duplicate_steps function for {student.CampusID}")
        logger.info(f"domain url : {domain_url}")
        logger.info(f"domain url : {domain_url+'QuickSearch.aspx?ForceCacheSearch=Y&amp;lname=&amp;fname='}")
        driver.get(domain_url+'QuickSearch.aspx?ForceCacheSearch=Y&amp;lname=&amp;fname=')
        wait = WebDriverWait(driver, 10)  # 10 seconds timeout
        element = wait.until(
            ec.visibility_of_element_located(
                (By.ID, config['ID_XPATH']['home_campus_id'])))

        # element = driver.find_element(By.ID, config['ID_XPATH']['home_campus_id'])
        # Click the element
        element.clear()
        element.send_keys(student.CampusID)
        # print("success1")
        # time.sleep(1)
        driver.find_element(By.ID, config['ID_XPATH']['search_button']).click()
        # print("success2")
        # time.sleep(1)
        # wait = WebDriverWait(driver, 10)  # 10 seconds timeout
        element = wait.until(
            ec.visibility_of_element_located(
                (By.XPATH, config['ID_XPATH']['href_select'])))
        # element = driver.find_element(By.XPATH, config['ID_XPATH']['href_select'])
        href_value = element.get_attribute("href")

        # Print the extracted href value
        # logger.info("Extracted href value:", href_value)
        match = re.search(r'\((.*?)\)', href_value)
        # extracted_value = ""
        if match:
            extracted_value = match.group(1)
            # logger.info("Extracted value inside parentheses:", extracted_value)
        else:
            logger.info("No value inside parentheses found.")
            raise Exception
        # logger.info("domain url+ additional : ", domain_url + 'BioInfo.aspx?Id='+extracted_value)
        driver.get(domain_url+'BioInfo.aspx?Id='+str(extracted_value))
        # time.sleep(1)
        # wait = WebDriverWait(driver, 10)  # 10 seconds timeout
        # wait = WebDriverWait(driver, 10)  # 10 seconds timeout
        element = wait.until(
            ec.visibility_of_element_located(
                (By.ID, config['ID_XPATH']['bio_link'])))
        # element = driver.find_element(By.ID, config['ID_XPATH']['bio_link'])
        element.click()
        # print("success4")
        # time.sleep(1)
        driver.find_element(By.ID, config['ID_XPATH']['edit_save_btn']).click()
        # print("success5")
        # time.sleep(1)
        # wait = WebDriverWait(driver, 10)  # 10 seconds timeout
        element = wait.until(
            ec.visibility_of_element_located(
                (By.ID, config['ID_XPATH']['home_campus_id'])))
        # driver.find_element(By.ID, config['ID_XPATH']['home_campus_id']).clear()
        # print("success6")
        element.clear()
        # time.sleep(1)
        Select(driver.find_element(By.ID, config['ID_XPATH']['database_status'])).select_by_value(" ")
        # print("success7")
        # time.sleep(1)
        driver.find_element(By.ID, config['ID_XPATH']['save_button']).click()
        # print("success8")
        # time.sleep(1)
        logger.info(f"successfully removed student inside duplicate_steps function for {student.CampusID}")
        driver.get(domain_url + 'AddNewIndividual.aspx')
        return True
    except Exception as e:
        logger.error("failed in duplicate_steps function :", e)
        return False
def duplicate_check(student, driver, domain_url, config, progress_bar):
    logger.info(f"inside duplicate_check function for {student.CampusID}")
    try:

        driver.get(domain_url + 'AddNewIndividual.aspx')  # with VPN
        # driver.get(domain_url + '/AddNewIndividual.aspx')  # without VPN
        wait = WebDriverWait(driver, 10)  # 10 seconds timeout
        element = wait.until(
            ec.visibility_of_element_located(
                (By.XPATH, config['ID_XPATH']['last_name'])))
        # Click the element
        element.send_keys(student.Surname)
        # time.sleep(1)
        driver.find_element(By.XPATH, config['ID_XPATH']['first_name']).send_keys(student.GivenName)
        # time.sleep(1)
        driver.find_element(By.XPATH, config['ID_XPATH']['birth_date']).send_keys(student.Birthdate)
        # time.sleep(1)
        driver.find_element(By.XPATH, config['ID_XPATH']['admissions_id']).send_keys(student.AdmissionsID)
        # time.sleep(1)
        driver.find_element(By.XPATH, config['ID_XPATH']['campus_id']).send_keys(student.CampusID)
        # time.sleep(1)
        Select(driver.find_element(By.ID, config['ID_XPATH']['department_id'])).select_by_visible_text(
            student.Department)
        # time.sleep(1)
        driver.find_element(By.ID, config['ID_XPATH']['continue_btn']).click()
        # time.sleep(1)
        logger.info(f"inside add individual page duplicate_check function for {student.CampusID}")
        # try:
        # time.sleep(1)
        try:
            driver.find_element(By.ID, config['ID_XPATH']['error_element'])
            # check_val = True
            logger.info(f"duplicate caught duplicate_check function for {student.CampusID}")
            raise Exception
            # Perform actions with the error element (if needed)
        except NoSuchElementException:
            logger.info(f"no exceptions caught duplicate_check function for {student.CampusID}")
            check_val = False
        except Exception as e:
            # time.sleep(1)
            logger.error(f"inside duplicate_exception duplicate_check function for {student.CampusID} and error {e}")
            status = adding_to_excel(student, driver, config, progress_bar)
            if not status:
                raise Exception
            # time.sleep(1)
            status = duplicate_steps(student, driver, domain_url, config)
            if not status:
                raise Exception
            else:
                check_val = True
        # time.sleep(1)
        status = AddIndividual(student, driver, check_val)
        return status
    except Exception as e:
        logger.error(e)
        return False


def process_student(std, url, config, driver, progress_bar):
    try:
        logger.info(f"starting process_student function for {std.CampusID}")
        try:
            status = duplicate_check(std, driver, url, config, progress_bar)
            if status:
                logger.info(f'Name: {std.GivenName}')
                logger.info(f'Birthdate: {std.Birthdate}')
                logger.info(f'Campus ID: {std.CampusID}')
                logger.info(f'Admissions ID: {std.AdmissionsID}')
                logger.info("Process completed successfully")
                progress_bar.success_count += 1
            else:
                logger.error(f"Process failed for student with ID {std.CampusID} and name {std.GivenName}.")
                progress_bar.failure_count += 1
        except Exception as e:
            logger.error(f"An error occurred for student with ID {std.CampusID}: {e}")
            progress_bar.failure_count += 1
    except Exception as e:
        logger.error(f"Process failed inside for process_student student : {std.CampusID} and error {e}")
        progress_bar.failure_count += 1
class Student:
    def __init__(self, AdmissionsID, CampusID, GivenName, Surname, Birthdate, Department, Template, BirthCountry,
                 Citizenship,
                 Gender, Level, Major, StartDate, SessionStartDate, EndDate, AcademicTerm, Tuition, LivingExpenses,
                 OtherDesc, OtherAmount,
                 PersonalFunds, SchoolFundsDesc, SchoolFunds, FundsFromAnother, FundsFromAnotherDesc, Remarks,
                 PermLine1, PermLine2,
                 PermCity, PermProvince, PermPostal, PermCountry, Email, phone_code, Phone):
        self.AdmissionsID = AdmissionsID
        self.CampusID = CampusID
        self.GivenName = GivenName
        self.Surname = Surname
        self.Birthdate = Birthdate
        self.Department = Department
        self.Template = Template
        self.BirthCountry = BirthCountry
        self.Citizenship = Citizenship
        self.Gender = Gender
        self.Level = Level
        self.Major = Major
        self.StartDate = StartDate
        self.SessionStartDate = SessionStartDate
        self.EndDate = EndDate
        self.AcademicTerm = AcademicTerm
        self.Tuition = Tuition
        self.LivingExpenses = LivingExpenses
        self.OtherDesc = OtherDesc
        self.OtherAmount = OtherAmount
        self.PersonalFunds = PersonalFunds
        self.SchoolFundsDesc = SchoolFundsDesc
        self.SchoolFunds = SchoolFunds
        self.FundsFromAnother = FundsFromAnother
        self.FundsFromAnotherDesc = FundsFromAnotherDesc
        self.Remarks = Remarks
        self.PermLine1 = PermLine1
        self.PermLine2 = PermLine2
        self.PermCity = PermCity
        self.PermProvince = PermProvince
        self.PermPostal = PermPostal
        self.PermCountry = PermCountry
        self.Email = Email
        self.phone_code = phone_code
        self.Phone = Phone

    def __str__(self):
        attributes = [
            f"Admissions ID: {self.AdmissionsID}",
            f"Campus ID: {self.CampusID}",
            f"Given Name: {self.GivenName}",
            f"Surname: {self.Surname}",
            f"Birthdate: {self.Birthdate}",
            f"Department: {self.Department}",
            f"Template: {self.Template}",
            f"Birth Country: {self.BirthCountry}",
            f"Citizenship: {self.Citizenship}",
            f"Gender: {self.Gender}",
            f"Level: {self.Level}",
            f"Major: {self.Major}",
            f"Start Date: {self.StartDate}",
            f"Session Start Date: {self.SessionStartDate}",
            f"End Date: {self.EndDate}",
            f"Academic Term: {self.AcademicTerm}",
            f"Tuition: {self.Tuition}",
            f"Living Expenses: {self.LivingExpenses}",
            f"Other Description: {self.OtherDesc}",
            f"Other Amount: {self.OtherAmount}",
            f"Personal Funds: {self.PersonalFunds}",
            f"School Funds Description: {self.SchoolFundsDesc}",
            f"School Funds: {self.SchoolFunds}",
            f"Funds From Another: {self.FundsFromAnother}",
            f"Funds From Another Description: {self.FundsFromAnotherDesc}",
            f"Remarks: {self.Remarks}",
            f"Permanent Address Line 1: {self.PermLine1}",
            f"Permanent Address Line 2: {self.PermLine2}",
            f"Permanent City: {self.PermCity}",
            f"Permanent Province: {self.PermProvince}",
            f"Permanent Postal Code: {self.PermPostal}",
            f"Permanent Country: {self.PermCountry}",
            f"Email: {self.Email}",
            f"Phone Country Code: {self.phone_code}",
            f"Phone: {self.Phone}"
        ]
        return '\n'.join(attributes)


def testing_main(url, driver, excel_file):
    # Load the Excel file into a DataFrame using pandas
    df = pd.read_excel("preProcessor/Duplicate.xlsx", engine='openpyxl')
    # Clear the DataFrame of any existing data (excluding the header)
    data = df.drop(df.index.to_list()[0:], axis=0)
    # Save the DataFrame (header row only) to the same Excel file
    with pd.ExcelWriter("preProcessor/Duplicate.xlsx", engine='openpyxl') as writer:
        data.to_excel(writer, index=False, sheet_name='Sheet1')
    # Print the DataFrame with only the header row
    # print(data)
    logger.info("Deletion success")
    try:
        logger.info("Main program started.")
        config = configparser.ConfigParser()
        config.read('preProcessor/config.ini')
        code_start_time = time.time()  # capturing the start time of the code execution
        # Load the Excel file
        logger.info(f"printing excel file name: {excel_file}")
        workbook = openpyxl.load_workbook(excel_file)
        # Select the specific sheet
        sheet = workbook['Export']  # Replace 'Sheet1' with the name of your sheet

        # Read the header row
        header_row = sheet[1]
        header_values = [cell.value for cell in header_row]
        num_rows_excluding_header = sheet.max_row - 1
        progress_bar = ProgressBar(num_rows_excluding_header)

        # Find the column indices for relevant fields
        AdmissionsID_index = header_values.index('AdmissionsID')
        CampusID_index = header_values.index('CampusID')
        GivenName_index = header_values.index('GivenName')
        Surname_index = header_values.index('Surname')
        Birthdate_index = header_values.index('Birthdate')
        Department_index = header_values.index('Department')
        Template_index = header_values.index('Template')
        BirthCountry_index = header_values.index('BirthCountry')
        Citizenship_index = header_values.index('Citizenship')
        Gender_index = header_values.index('Gender')
        Level_index = header_values.index('Level')
        Major_index = header_values.index('Major')
        StartDate_index = header_values.index('StartDate')
        SessionStartDate_index = header_values.index('SessionStartDate')
        EndDate_index = header_values.index('EndDate')
        AcademicTerm_index = header_values.index('AcademicTerm')
        Tuition_index = header_values.index('Tuition')
        LivingExpenses_index = header_values.index('LivingExpenses')
        OtherDesc_index = header_values.index('OtherDesc')
        OtherAmount_index = header_values.index('OtherAmount')
        PersonalFunds_index = header_values.index('PersonalFunds')
        SchoolFundsDesc_index = header_values.index('SchoolFundsDesc')
        SchoolFunds_index = header_values.index('SchoolFunds')
        FundsFromAnother_index = header_values.index('FundsFromAnother')
        FundsFromAnotherDesc_index = header_values.index('FundsFromAnotherDesc')
        Remarks_index = header_values.index('Remarks')
        PermLine1_index = header_values.index('PermLine1')
        PermLine2_index = header_values.index('PermLine2')
        PermCity_index = header_values.index('PermCity')
        PermProvince_index = header_values.index('PermProvince')
        PermPostal_index = header_values.index('PermPostal')
        PermCountry_index = header_values.index('PermCountry')
        Email_index = header_values.index('Email')
        Phone_index = header_values.index('Phone')

        # Read the data rows
        students = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            AdmissionsID = '' if pd.isna(row[AdmissionsID_index]) else row[AdmissionsID_index]
            CampusID = '' if pd.isna(row[CampusID_index]) else row[CampusID_index]
            GivenName = '' if pd.isna(row[GivenName_index]) else row[GivenName_index]
            Surname = '' if pd.isna(row[Surname_index]) else row[Surname_index]
            Birthdate = '' if pd.isna(row[Birthdate_index]) else row[Birthdate_index]
            Department = '' if pd.isna(row[Department_index]) else row[Department_index]
            Template = '' if pd.isna(row[Template_index]) else row[Template_index]
            BirthCountry = '' if pd.isna(row[BirthCountry_index]) else row[BirthCountry_index]
            Citizenship = '' if pd.isna(row[Citizenship_index]) else row[Citizenship_index]
            Gender = '' if pd.isna(row[Gender_index]) else row[Gender_index]
            Level = '' if pd.isna(row[Level_index]) else row[Level_index]
            Major = '' if pd.isna(row[Major_index]) else row[Major_index]
            StartDate = row[StartDate_index]
            SessionStartDate = row[SessionStartDate_index]
            EndDate = row[EndDate_index]
            AcademicTerm = '' if pd.isna(row[AcademicTerm_index]) else row[AcademicTerm_index]
            Tuition = '' if pd.isna(row[Tuition_index]) else row[Tuition_index]
            LivingExpenses = row[LivingExpenses_index]
            OtherDesc = row[OtherDesc_index]
            OtherAmount = row[OtherAmount_index]
            PersonalFunds = row[PersonalFunds_index]
            SchoolFundsDesc = '' if pd.isna(row[SchoolFundsDesc_index]) else row[SchoolFundsDesc_index]
            SchoolFunds = '' if pd.isna(row[SchoolFunds_index]) else row[SchoolFunds_index]
            FundsFromAnother = '' if pd.isna(row[FundsFromAnother_index]) else row[FundsFromAnother_index]
            FundsFromAnotherDesc = '' if pd.isna(row[FundsFromAnotherDesc_index]) else row[FundsFromAnotherDesc_index]
            Remarks = row[Remarks_index]
            PermLine1 = '' if pd.isna(row[PermLine1_index]) else row[PermLine1_index]
            PermLine2 = '' if pd.isna(row[PermLine2_index]) else row[PermLine2_index]
            PermCity = '' if pd.isna(row[PermCity_index]) else row[PermCity_index]
            PermProvince = '' if pd.isna(row[PermProvince_index]) else row[PermProvince_index]
            PermPostal = '' if pd.isna(row[PermPostal_index]) else row[PermPostal_index]
            PermCountry = '' if pd.isna(row[PermCountry_index]) else row[PermCountry_index]
            Email = '' if pd.isna(row[Email_index]) else row[Email_index]
            phone_number = '' if pd.isna(row[Phone_index]) else str(row[Phone_index])
            phone_code = ''
            if phone_number.startswith('+'):
                parts = phone_number.split(' ', 1)
                if len(parts) > 1:
                    phone_code = parts[0].replace('+', '')
                    phone_number = parts[1].replace(' ', '').replace('-', '')
            if any(val != '' for val in
                   [AdmissionsID, CampusID, GivenName, Surname, Birthdate, Department, Template, BirthCountry,
                    Citizenship,
                    Gender, Level, Major, StartDate, SessionStartDate, EndDate, AcademicTerm, Tuition, LivingExpenses,
                    OtherDesc, OtherAmount, PersonalFunds, SchoolFundsDesc, SchoolFunds, FundsFromAnother,
                    FundsFromAnotherDesc,
                    Remarks, PermLine1, PermLine2, PermCity, PermProvince, PermPostal, PermCountry, Email, phone_code,
                    phone_number]):
                student = Student(AdmissionsID, CampusID, GivenName, Surname, Birthdate, Department, Template,
                                  BirthCountry, Citizenship,
                                  Gender, Level, Major, StartDate, SessionStartDate, EndDate, AcademicTerm, Tuition,
                                  LivingExpenses, OtherDesc, OtherAmount,
                                  PersonalFunds, SchoolFundsDesc, SchoolFunds, FundsFromAnother, FundsFromAnotherDesc,
                                  Remarks, PermLine1, PermLine2,
                                  PermCity, PermProvince, PermPostal, PermCountry, Email, phone_code, phone_number)
                students.append(student)

        # Access each student's birthdate and campus ID
        logger.info("Starting records iteration")
        # with ThreadPoolExecutor(max_workers=15) as executor:
        #     # Submit each student for processing concurrently
        #     executor.map(process_student, students, url, config)
        for index, student in enumerate(students):
            logger.info(f"Index No : {index}, student ID : {student.CampusID}")
            process_student(student, url, config, driver, progress_bar)
            progress_bar.processed_count = index + 1
            progressBar_value = math.floor((progress_bar.processed_count/progress_bar.max_count)*6)
            logger.info(f"percentage completed: {progressBar_value}")
            socketio.emit('preProcessor', progressBar_value)
            logger.info(progress_bar.__str__())

        code_end_time = time.time()  # capturing the end time of the code execution
        total_time = code_end_time - code_start_time  # calculating the total execution time
        logger.info("Total execution time: {:.2f} seconds".format(total_time))  # logging the total execution time
        logger.info("Main program finished.")


        if progress_bar.deferral_count > 0 and progress_bar.max_count == progress_bar.success_count:
            logger.info(f"Deferral cases in this batch run.")
            message = "Partial success"
            return True, message
        elif progress_bar.max_count == progress_bar.success_count:
            logger.info(f"No failure cases in this batch run.")
            message = "Success"
            return True, message
        elif progress_bar.failure_count == progress_bar.max_count:
            logger.info(f"All cases in this batch run failed.")
            message = "Failed"
            return True, message
        else:
            logger.info(f"mixed cases i.e. success and failure both.")
            message = "Failed"
            return True, message
            #  need to handle this mixed cases response in main.py and front end to show either in logs or after run.
    except Exception as e:
        logger.error(f"An error occurred in testing single.py testing_main function: {e}")
        return False, "Failed"